import numpy as np
from scipy import random
from multiprocessing import Pool
from openpyxl import Workbook

# # set print options for large arrays
np.set_printoptions(threshold=np.inf, precision=2, linewidth=np.inf)

m = 5  # number of attributes
n = 10  # number of instances

# initialize mean vector
mean = []
for i in range(m):
    mean.append(random.randint(0, 100))
mean = np.array(mean)
print('Mean: \n', mean)

# initialize covariance matrix (must be positive semi-definite)
A = random.rand(m, m)
cov = np.dot(A, A.transpose())
print('Covariance Matrix: \n', cov)


def f():  # no argument
    return np.random.multivariate_normal(mean, cov, 1)


def throw_away_function(_):
    return f()

# large_array = Pool(1).map(throw_away_function, range(n))
# print(np.array(large_array))
# print(len(large_array))

wb = Workbook()
ws = wb.create_sheet('Data')


for i in range(n):
    row = np.random.multivariate_normal(mean, cov, 1)[0]
    print(row)
    for j in range(m):
        ws.cell(column=j + 1, row=i + 1, value=row[j])

wb.save(filename='../assets/bla_excel.xlsx')
