# only create file with unlabeled and normally distributed instances ones

from scipy.sparse import csr_matrix
from scipy.sparse.csgraph import minimum_spanning_tree
import numpy as np
import matplotlib.pyplot as plt
from random import randint, random
from openpyxl import Workbook
import csv

# -------- Parameters --------
m = 10  # number of features
n = 1000  # number of instances
b = 0.5  # the desired complexity, defined by the length of the class boundary, b ∈[0,1].
l = [0, 1]  # manually define a set of possible labels
# -------- End Parameters --------

# TODO switch to csv
# Excel Workbook
wb = Workbook()
excel_sheet = wb.create_sheet(title='Data')
# Set Header in Excel
for i in range(1, m + 1, 1):
    column_title = "Feature %s" % i
    excel_sheet.cell(column=i, row=1, value=column_title)

for i in range(0, m, 1):
    # set distribution for this feature
    mu, sigma = random() * 100, random() * 100  # TODO distribution for mu and sigma itself ?
    print('Mu: %dm, Sigma: %s' % (mu, sigma))
    s = np.random.normal(mu, sigma, n)
    print(s)
    print()
    with open('some.csv', 'w', newline='') as f:
        writer = csv.writer(f)
        writer.writerows(np.array(s))
    for j in range(0, n, 1):
        excel_sheet.cell(column=i + 1, row=j + 2, value=int(s[j]))


    # wb.save(filename='../assets/Data.xlsx')

