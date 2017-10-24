from openpyxl import Workbook

#  Excel
excel_file = Workbook()
excel_sheet = excel_file.create_sheet(title='m=' + str(m))
excel_sheet.cell(column=1, row=1, value='Nr.')
for i in range(len(complexity_measures)):
    excel_sheet.cell(column=(i+2), row=1, value=complexity_measures[i])