import os
import pandas as pd
from openpyxl import Workbook

from generator.functions import create_dataset_and_or_mst


def evaluate(individual, mst_edges, n):
    # 1. Store the nodes of the spanning tree with different class.
    nodes = [-1] * n
    for edge in mst_edges:
        if individual[edge[0]] != individual[edge[1]]:
            nodes[edge[0]] = 0
            nodes[edge[1]] = 0

    # 2. Compute the number of nodes of the spanning tree with different class.
    different = 0
    for i in range(n):
        if nodes[i] == 0:
            different += 1
    # print('different:', different)
    return different / n
#
#
# excel_file = Workbook()
#
# # complexities_dict = {1: 0.2, 2: 0.4, 3: 0.6, 4: 0.8}
#
# sub_complexity = [0.2, 0.4, 0.6, 0.8]
# meta_complexity = [0.001, 0.01, 0.1]
#
# # iterate through all data sets
# for iter in range(1):
#     excel_sheet = excel_file.create_sheet(title=str((iter + 1)))
#     # header
#     excel_sheet.cell(column=1, row=1, value='Csub')
#     excel_sheet.cell(column=2, row=1, value='Cmeta')
#     excel_sheet.cell(column=3, row=1, value='Ccomplete')
#     print('ITERATION')
#
#     for sub in sub_complexity:
#         if not os.path.exists('../assets_/complexity_%r' % sub):
#                 os.makedirs('../assets_/complexity_%r' % sub)
#
#     row = 2
#     for sub in sub_complexity:
#         excel_sheet.cell(column=1, row=row, value=sub)
#         for meta in sub_complexity:
#             excel_sheet.cell(column=2, row=row, value=meta)
#             complete_data = pd.read_csv(
#                 filepath_or_buffer='../assets/complexity_%r/data_%r_meta_%r.csv' % (sub, (iter + 1), meta))
#             complete_data_ = pd.read_csv(
#                 filepath_or_buffer='../assets/complexity_%r/data_%r_meta_%r.csv' % (sub, (iter + 1), meta))
#
#             # print('../assets/complexity_%r/data_%r.csv' % (meta, (iter + 1)))
#
#             # print(complete_data['label'])
#             complete_mst = create_dataset_and_or_mst(data=complete_data_,
#                                                      path='../assets/complexity_%r/data_%r_meta_%r.csv' % (
#                                                      sub, (iter + 1), meta))
#             print(complete_data['label'])
#             C_complete = evaluate(individual=complete_data['label'], mst_edges=complete_mst, n=1000)
#
#             # print(C_complete)
#             excel_sheet.cell(column=3, row=row, value=C_complete)
#
#             excel_file.save(filename='results.xlsx')
#
#             row += 1
